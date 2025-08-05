# import pandas as pd
# import os
# from datetime import datetime, timedelta
# from utils import get_month_folder, get_date_str, is_weekend, load_log, save_log
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# MASTER_COLUMNS = ["Office", "Emp Code", "Device ID", "Name"]
# WEEK_TOTAL_PREFIX = "Week"

# RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
# WEEKEND_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# # Original dictionary for TVM
# DEVICE_TO_EMP = {
#     214: "181",
#     679: "589",
#     808: "691",
#     740: "629",
#     970: "I157",
#     227: "194",
#     612: "524",
#     1004: "I176",
#     61: "71",
#     777: "663",
#     587: "501",
#     692: "624",
#     8: "72",
#     491: "408",
#     800: "684",
#     65: "83",
#     267: "213",
#     791: "677",
#     1002: "829",
#     995: "I172",
#     914: "782",
#     "": "347",
#     796: "236",
#     489: "406",
#     973: "819",
#     579: "492",
#     940: "797",
#     110: "124",
#     1029: "",
#     335: "285",
#     1031: "",
#     980: "827",
#     985: "I162",
#     913: "781",
#     461: "380",
#     1015: "837",
#     896: "767",
#     861: "173",
#     990: "I167",
#     220: "187",
#     1028: "847",
#     448: "369",
#     352: "292",
#     810: "693",
#     613: "525",
#     767: "655",
#     986: "I163",
#     290: "234",
#     997: "I174",
#     280: "229",
#     961: "",
#     522: "435",
#     5: "115",
#     804: "687",
#     811: "694",
#     827: "711",
#     265: "203",
#     24: "24",
#     919: "786",
#     850: "161",
#     987: "I164",
#     960: "814",
#     1006: "",
#     996: "I173",
#     988: "I165",
#     984: "I161",
#     981: "I158",
#     614: "526",
#     812: "695",
#     982: "I159",
#     979: "826",
#     648: "559",
#     1001: "828",
#     625: "582",
#     805: "688",
#     847: "",
#     288: "233",
#     275: "223",
#     753: "643",
#     657: "567",
#     159: "166",
#     54: "80",
#     830: "714",
#     "": "638",
#     857: "CO88",
#     571: "485",
#     610: "522",
#     955: "CO120",
#     963: "851",
#     968: "817",
#     595: "507",
#     874: "744",
#     707: "605",
#     769: "109",
#     13: "73",
#     195: "167",
#     942: "799",
#     408: "498",
#     1027: "846",
#     393: "",
#     596: "508",
#     834: "719",
#     775: "661",
#     373: "296",
#     1032: "",
#     32: "107",
#     949: "363",
#     938: "",
#     771: "657",
#     998: "I175",
#     946: "803",
#     708: "716",
#     701: "30",
#     994: "I171",
#     437: "353",
#     307: "246",
#     956: "810",
#     954: "809",
#     183: "163",
#     846: "726",
#     243: "239",
#     943: "800",
#     615: "527",
#     823: "707",
#     828: "712",
#     167: "154",
#     840: "754",
#     517: "430",
#     891: "623",
#     921: "788",
#     1008: "832",
#     325: "257",
#     308: "247",
#     1036: "850",
#     875: "745",
#     871: "741",
#     636: "548",
#     411: "326",
#     992: "I169",
#     268: "216",
#     231: "193",
#     400: "460",
#     1017: "838",
#     495: "412",
#     724: "176",
#     1012: "833",
#     535: "448",
#     757: "645",
#     742: "",
#     536: "449",
#     1023: "843",
#     983: "I160",
#     617: "529",
#     601: "513",
#     545: "458",
#     602: "514",
#     680: "590",
#     345: "268",
#     1022: "587",
#     70: "81",
#     884: "755",
#     441: "357",
#     749: "635",
#     476: "394",
#     836: "721",
#     993: "I170",
#     907: "776",
#     235: "189",
#     843: "CO80",
#     941: "798",
#     975: "821",
#     755: "642",
#     1037: "",
#     364: "68",
#     910: "822",
#     344: "298",
#     520: "433",
#     609: "521",
#     180: "162",
#     342: "142",
#     1005: "",
#     1009: "164",
#     887: "760",
#     974: "820",
#     423: "339",
#     291: "235",
#     1020: "",
#     262: "219",
#     1030: "",
#     999: "",
#     416: "332",
#     790: "676",
#     789: "675",
#     991: "I168",
#     966: "815",
#     659: "569",
#     817: "700",
#     554: "468",
#     851: "729",
#     918: "785",
#     390: "312",
#     420: "336",
#     746: "704",
#     189: "185",
#     1019: "840",
#     211: "114",
#     854: "",
#     868: "756",
#     7: "7",
#     989: "I166",
#     202: "170",
#     574: "487"
# }

# # Kochi dictionary
# KOCHI_DEVICE_TO_EMP = {
#     14: "501",
#     15: "408",
#     28: "677",
#     29: "347",
#     9: "492",
#     21: "205",
#     24: "507",
#     23: "719",
#     19: "832",
#     16: "548",
#     7: "843",
#     27: "326",
#     18: "838",
#     1048: "268",
#     22: "587",
#     17: "821",
#     26: "468",
#     1019: "840",
#     458: "458",
#     25: "700",
#     13: "I165",
#     1039: "",
#     1040: "",
#     1041: "",
#     1042: "",
#     1043: "",
#     1044: ""
# }

# def load_or_create_master(month_folder, emp_list):
#     master_path = os.path.join(month_folder, "master.xlsx")
#     if os.path.exists(master_path):
#         return pd.read_excel(master_path)
#     else:
#         df = pd.DataFrame(emp_list, columns=MASTER_COLUMNS)
#         return df

# def time_str_to_hours(t):
#     try:
#         if isinstance(t, str) and ":" in t:
#             h, m = map(int, t.strip().split(":"))
#             return h + m / 60
#         elif isinstance(t, (int, float)):
#             return float(t)
#         else:
#             return 0
#     except:
#         return 0

# def fill_missing_days(master):
#     all_dates = [col for col in master.columns if col not in MASTER_COLUMNS and not col.startswith(WEEK_TOTAL_PREFIX) and col != "Month Total"]
#     all_dates_sorted = sorted(all_dates, key=lambda x: datetime.strptime(x, "%d-%b"))
#     existing = set(all_dates_sorted)
#     for i in range(len(all_dates_sorted) - 1):
#         d1 = datetime.strptime(all_dates_sorted[i], "%d-%b")
#         d2 = datetime.strptime(all_dates_sorted[i + 1], "%d-%b")
#         for day in range(1, (d2 - d1).days):
#             label = (d1 + timedelta(days=day)).strftime("%d-%b")
#             if label not in existing:
#                 master[label] = "-"
#     return master

# # def insert_week_totals(master):
# #     date_cols = [col for col in master.columns if col not in MASTER_COLUMNS and not col.startswith(WEEK_TOTAL_PREFIX) and col != "Month Total"]
# #     date_objs = [(datetime.strptime(col, "%d-%b"), col) for col in date_cols]
# #     date_objs.sort()
# #     result = master[MASTER_COLUMNS].copy()
# #     week_number = 1
# #     temp_cols = []
# #     for date_obj, col in date_objs:
# #         result[col] = master[col]
# #         temp_cols.append(col)
# #         if date_obj.weekday() == 4:
# #             week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
# #             result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
# #             week_number += 1
# #             temp_cols = []
# #     if temp_cols:
# #         week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
# #         result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
# #     month_cols = [c for c in result.columns if c not in MASTER_COLUMNS and not c.startswith(WEEK_TOTAL_PREFIX) and c != "Month Total"]
# #     result["Month Total"] = result[month_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
# #     return result
# def insert_week_totals(master):
#     date_cols = [col for col in master.columns if col not in MASTER_COLUMNS and not col.startswith(WEEK_TOTAL_PREFIX) and col != "Month Total"]
#     date_objs = [(datetime.strptime(col, "%d-%b"), col) for col in date_cols]
#     date_objs.sort()

#     result = master[MASTER_COLUMNS].copy()
#     week_number = 1
#     temp_cols = []
#     workday_count = 0

#     for date_obj, col in date_objs:
#         result[col] = master[col]
#         temp_cols.append(col)
#         if date_obj.weekday() < 5:  # Only count Mon–Fri
#             workday_count += 1
#         if workday_count == 5:  # After 5 weekdays, insert Week Total
#             week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
#             result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
#             week_number += 1
#             temp_cols = []
#             workday_count = 0

#     # For any remaining days (less than 5), insert one more week total
#     if temp_cols:
#         week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
#         result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)

#     # Add Month Total
#     month_cols = [c for c in result.columns if c not in MASTER_COLUMNS and not c.startswith(WEEK_TOTAL_PREFIX) and c != "Month Total"]
#     result["Month Total"] = result[month_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
#     return result

# def apply_excel_formatting(file_path, year):
#     wb = load_workbook(file_path)
#     ws = wb.active
#     header = [cell.value for cell in ws[1]]
#     for col_index, heading in enumerate(header, 1):
#         try:
#             date_obj = datetime.strptime(f"{heading}-{year}", "%d-%b-%Y")
#             if is_weekend(date_obj):
#                 for row in range(1, ws.max_row + 1):
#                     ws.cell(row=row, column=col_index).fill = WEEKEND_FILL
#             else:
#                 for row in range(2, ws.max_row + 1):
#                     val = ws.cell(row=row, column=col_index).value
#                     if val == "Late":
#                         ws.cell(row=row, column=col_index).fill = RED_FILL
#                     elif val == "-":
#                         ws.cell(row=row, column=col_index).fill = YELLOW_FILL
#         except:
#             continue
#     wb.save(file_path)

# def process_attendance(trivandrum_file, kochi_file):
#     today = get_date_str(trivandrum_file)
#     if is_weekend(today):
#         return {"status": "skipped", "reason": "Weekend"}

#     log = load_log()
#     if today.strftime("%Y-%m-%d") in log.get("processed", []):
#         return {"status": "skipped", "reason": "Already processed"}

#     try:
#         trivandrum_df = pd.read_excel(trivandrum_file, skiprows=9)
#         kochi_df = pd.read_excel(kochi_file, skiprows=9)
#     except Exception as e:
#         return {"status": "error", "reason": f"Failed to read Excel: {str(e)}"}

#     trivandrum_df.columns = [col.strip() for col in trivandrum_df.columns]
#     kochi_df.columns = [col.strip() for col in kochi_df.columns]

#     trivandrum_df = trivandrum_df.fillna(0)
#     kochi_df = kochi_df.fillna(0)

#     daily_df = pd.concat([trivandrum_df, kochi_df], ignore_index=True)

#     attendance = {row["Emp Code"]: time_str_to_hours(row["In Duration (In Hrs)"]) for _, row in daily_df.iterrows()}

#     month_folder = get_month_folder(today)
#     os.makedirs(month_folder, exist_ok=True)

#     emp_list = []

#     for _, row in trivandrum_df.iterrows():
#         device_id = row["Emp Code"]
#         emp_code = DEVICE_TO_EMP.get(device_id, "UNKNOWN")
#         emp_list.append(("TVM", emp_code, device_id, row["Name"]))

#     for _, row in kochi_df.iterrows():
#         device_id = row["Emp Code"]
#         emp_code = KOCHI_DEVICE_TO_EMP.get(device_id, "UNKNOWN")
#         emp_list.append(("Kochi", emp_code, device_id, row["Name"]))

#     master = load_or_create_master(month_folder, emp_list)

#     existing_device_ids = set(master["Device ID"])
#     for office, emp_code, device_id, name in emp_list:
#         if device_id not in existing_device_ids:
#             new_row = {"Office": office, "Emp Code": emp_code, "Device ID": device_id, "Name": name}
#             for col in master.columns:
#                 if col not in MASTER_COLUMNS:
#                     new_row[col] = "-" if all(val == "-" for val in master[col] if isinstance(val, str)) else "No"
#             master = pd.concat([master, pd.DataFrame([new_row])], ignore_index=True)

#     date_col = today.strftime("%d-%b")
#     if date_col not in master.columns:
#         master[date_col] = "No"

#     for idx, row in master.iterrows():
#         device_id = row["Device ID"]
#         duration = attendance.get(device_id, 0)
#         if device_id in attendance:
#             if duration >= 5:
#                 master.at[idx, date_col] = "Yes"
#             elif duration > 0:
#                 master.at[idx, date_col] = "Late"
#             else:
#                 master.at[idx, date_col] = "No"

#     master = fill_missing_days(master)
#     master = insert_week_totals(master)

#     master_path = os.path.join(month_folder, "master.xlsx")
#     if os.path.exists(master_path):
#         os.remove(master_path)
#     master.to_excel(master_path, index=False)
#     apply_excel_formatting(master_path, today.year)

#     log.setdefault("processed", []).append(today.strftime("%Y-%m-%d"))
#     save_log(log)

#     return {"status": "success", "date": today.strftime("%Y-%m-%d")}
import pandas as pd
import os
from datetime import datetime, timedelta
from utils import get_month_folder, get_date_str, is_weekend, load_log, save_log
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
MASTER_COLUMNS = ["Office", "Emp Code", "Device ID", "Name"]
WEEK_TOTAL_PREFIX = "Week"
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Device-to-employee mappings
DEVICE_TO_EMP = {
    214: "181",
    679: "589",
    808: "691",
    740: "629",
    970: "I157",
    227: "194",
    612: "524",
    1004: "I176",
    61: "71",
    777: "663",
    587: "501",
    692: "624",
    8: "72",
    491: "408",
    800: "684",
    65: "83",
    267: "213",
    791: "677",
    1002: "829",
    995: "I172",
    914: "782",
    "": "347",
    796: "236",
    489: "406",
    973: "819",
    579: "492",
    940: "797",
    110: "124",
    1029: "",
    335: "285",
    1031: "",
    980: "827",
    985: "I162",
    913: "781",
    461: "380",
    1015: "837",
    896: "767",
    861: "173",
    990: "I167",
    220: "187",
    1028: "847",
    448: "369",
    352: "292",
    810: "693",
    613: "525",
    767: "655",
    986: "I163",
    290: "234",
    997: "I174",
    280: "229",
    961: "",
    522: "435",
    5: "115",
    804: "687",
    811: "694",
    827: "711",
    265: "203",
    24: "24",
    919: "786",
    850: "161",
    987: "I164",
    960: "814",
    1006: "",
    996: "I173",
    988: "I165",
    984: "I161",
    981: "I158",
    614: "526",
    812: "695",
    982: "I159",
    979: "826",
    648: "559",
    1001: "828",
    625: "582",
    805: "688",
    847: "",
    288: "233",
    275: "223",
    753: "643",
    657: "567",
    159: "166",
    54: "80",
    830: "714",
    "": "638",
    857: "CO88",
    571: "485",
    610: "522",
    955: "CO120",
    963: "851",
    968: "817",
    595: "507",
    874: "744",
    707: "605",
    769: "109",
    13: "73",
    195: "167",
    942: "799",
    408: "498",
    1027: "846",
    393: "",
    596: "508",
    834: "719",
    775: "661",
    373: "296",
    1032: "",
    32: "107",
    949: "363",
    938: "",
    771: "657",
    998: "I175",
    946: "803",
    708: "716",
    701: "30",
    994: "I171",
    437: "353",
    307: "246",
    956: "810",
    954: "809",
    183: "163",
    846: "726",
    243: "239",
    943: "800",
    615: "527",
    823: "707",
    828: "712",
    167: "154",
    840: "754",
    517: "430",
    891: "623",
    921: "788",
    1008: "832",
    325: "257",
    308: "247",
    1036: "850",
    875: "745",
    871: "741",
    636: "548",
    411: "326",
    992: "I169",
    268: "216",
    231: "193",
    400: "460",
    1017: "838",
    495: "412",
    724: "176",
    1012: "833",
    535: "448",
    757: "645",
    742: "",
    536: "449",
    1023: "843",
    983: "I160",
    617: "529",
    601: "513",
    545: "458",
    602: "514",
    680: "590",
    345: "268",
    1022: "587",
    70: "81",
    884: "755",
    441: "357",
    749: "635",
    476: "394",
    836: "721",
    993: "I170",
    907: "776",
    235: "189",
    843: "CO80",
    941: "798",
    975: "821",
    755: "642",
    1037: "",
    364: "68",
    910: "822",
    344: "298",
    520: "433",
    609: "521",
    180: "162",
    342: "142",
    1005: "",
    1009: "164",
    887: "760",
    974: "820",
    423: "339",
    291: "235",
    1020: "",
    262: "219",
    1030: "",
    999: "",
    416: "332",
    790: "676",
    789: "675",
    991: "I168",
    966: "815",
    659: "569",
    817: "700",
    554: "468",
    851: "729",
    918: "785",
    390: "312",
    420: "336",
    746: "704",
    189: "185",
    1019: "840",
    211: "114",
    854: "",
    868: "756",
    7: "7",
    989: "I166",
    202: "170",
    574: "487"
}

# Kochi dictionary
KOCHI_DEVICE_TO_EMP = {
    14: "501",
    15: "408",
    28: "677",
    29: "347",
    9: "492",
    21: "205",
    24: "507",
    23: "719",
    19: "832",
    16: "548",
    7: "843",
    27: "326",
    18: "838",
    1048: "268",
    22: "587",
    17: "821",
    26: "468",
    1019: "840",
    458: "458",
    25: "700",
    13: "I165",
    1039: "",
    1040: "",
    1041: "",
    1042: "",
    1043: "",
    1044: ""
}
# Load or create master file
def load_or_create_master(month_folder, emp_list):
    master_path = os.path.join(month_folder, "master.xlsx")
    if os.path.exists(master_path):
        return pd.read_excel(master_path)
    else:
        return pd.DataFrame(emp_list, columns=MASTER_COLUMNS)

# Convert time string to float hours
def time_str_to_hours(t):
    try:
        if isinstance(t, str) and ":" in t:
            h, m = map(int, t.strip().split(":"))
            return h + m / 60
        elif isinstance(t, (int, float)):
            return float(t)
        else:
            return 0
    except:
        return 0

# Fill missing dates in master
def fill_missing_days(master):
    all_dates = [col for col in master.columns if col not in MASTER_COLUMNS and not col.startswith(WEEK_TOTAL_PREFIX) and col != "Month Total"]
    all_dates_sorted = sorted(all_dates, key=lambda x: datetime.strptime(x, "%d-%b"))
    existing = set(all_dates_sorted)
    for i in range(len(all_dates_sorted) - 1):
        d1 = datetime.strptime(all_dates_sorted[i], "%d-%b")
        d2 = datetime.strptime(all_dates_sorted[i + 1], "%d-%b")
        for day in range(1, (d2 - d1).days):
            label = (d1 + timedelta(days=day)).strftime("%d-%b")
            if label not in existing:
                master[label] = "-"
    return master

# Insert weekly and monthly totals
def insert_week_totals(master):
    date_cols = [col for col in master.columns if col not in MASTER_COLUMNS and not col.startswith(WEEK_TOTAL_PREFIX) and col != "Month Total"]
    date_objs = [(datetime.strptime(col, "%d-%b"), col) for col in date_cols]
    date_objs.sort()

    result = master[MASTER_COLUMNS].copy()
    week_number = 1
    temp_cols = []
    workday_count = 0

    for date_obj, col in date_objs:
        result[col] = master[col]
        temp_cols.append(col)
        if date_obj.weekday() < 5:  # Mon–Fri
            workday_count += 1
        if workday_count == 5:
            week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
            result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
            week_number += 1
            temp_cols = []
            workday_count = 0

    if temp_cols:
        week_col = f"{WEEK_TOTAL_PREFIX} {week_number}"
        result[week_col] = master[temp_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)

    month_cols = [c for c in result.columns if c not in MASTER_COLUMNS and not c.startswith(WEEK_TOTAL_PREFIX) and c != "Month Total"]
    result["Month Total"] = result[month_cols].apply(lambda row: sum(x == "Yes" for x in row), axis=1)
    return result

# Apply color formatting to Excel
def apply_excel_formatting(file_path, year):
    wb = load_workbook(file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    for col_index, heading in enumerate(header, 1):
        try:
            date_obj = datetime.strptime(f"{heading}-{year}", "%d-%b-%Y")
            if is_weekend(date_obj):
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col_index).fill = WEEKEND_FILL
            else:
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=col_index).value
                    if val == "Late":
                        ws.cell(row=row, column=col_index).fill = RED_FILL
                    elif val == "-":
                        ws.cell(row=row, column=col_index).fill = YELLOW_FILL
        except:
            continue
    wb.save(file_path)

# Main attendance processing
def process_attendance(trivandrum_file, kochi_file):
    today = get_date_str(trivandrum_file)
    if is_weekend(today):
        return {"status": "skipped", "reason": "Weekend"}

    log = load_log()
    if today.strftime("%Y-%m-%d") in log.get("processed", []):
        return {"status": "skipped", "reason": "Already processed"}

    try:
        trivandrum_df = pd.read_excel(trivandrum_file, skiprows=9)
        kochi_df = pd.read_excel(kochi_file, skiprows=9)
    except Exception as e:
        return {"status": "error", "reason": f"Failed to read Excel: {str(e)}"}

    trivandrum_df.columns = [col.strip() for col in trivandrum_df.columns]
    kochi_df.columns = [col.strip() for col in kochi_df.columns]
    trivandrum_df = trivandrum_df.fillna(0)
    kochi_df = kochi_df.fillna(0)

    daily_df = pd.concat([trivandrum_df, kochi_df], ignore_index=True)
    attendance = {row["Emp Code"]: time_str_to_hours(row["In Duration (In Hrs)"]) for _, row in daily_df.iterrows()}

    month_folder = get_month_folder(today)
    os.makedirs(month_folder, exist_ok=True)

    emp_list = []

    for _, row in trivandrum_df.iterrows():
        device_id = row["Emp Code"]
        emp_code = DEVICE_TO_EMP.get(device_id, "UNKNOWN")
        emp_list.append(("TVM", emp_code, device_id, row["Name"]))

    for _, row in kochi_df.iterrows():
        device_id = row["Emp Code"]
        emp_code = KOCHI_DEVICE_TO_EMP.get(device_id, "UNKNOWN")
        emp_list.append(("Kochi", emp_code, device_id, row["Name"]))

    master = load_or_create_master(month_folder, emp_list)

    # Add missing employees
    existing_device_ids = set(master["Device ID"])
    for office, emp_code, device_id, name in emp_list:
        if device_id not in existing_device_ids:
            new_row = {"Office": office, "Emp Code": emp_code, "Device ID": device_id, "Name": name}
            for col in master.columns:
                if col not in MASTER_COLUMNS:
                    new_row[col] = "-" if all(val == "-" for val in master[col] if isinstance(val, str)) else "No"
            master = pd.concat([master, pd.DataFrame([new_row])], ignore_index=True)

    # Add today's column if missing
    date_col = today.strftime("%d-%b")
    if date_col not in master.columns:
        master[date_col] = "No"

    # Mark attendance
    for idx, row in master.iterrows():
        device_id = row["Device ID"]
        duration = attendance.get(device_id, 0)
        if device_id in attendance:
            if duration >= 5:
                master.at[idx, date_col] = "Yes"
            elif duration > 0:
                master.at[idx, date_col] = "Late"
            else:
                master.at[idx, date_col] = "No"

    master = fill_missing_days(master)
    master = insert_week_totals(master)

    # Save master file
    master_path = os.path.join(month_folder, "master.xlsx")
    if os.path.exists(master_path):
        os.remove(master_path)
    master.to_excel(master_path, index=False)
    apply_excel_formatting(master_path, today.year)

    # Log processed date
    log.setdefault("processed", []).append(today.strftime("%Y-%m-%d"))
    save_log(log)

    return {"status": "success", "date": today.strftime("%Y-%m-%d")}
